import argparse
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import quote_sheetname
import requests
from requests.auth import HTTPBasicAuth
from dotenv import load_dotenv
import os


def create_users_from_file(file, nextcloudURL, overwriteUsers, app_user, app_pw):

    headers = {
        'OCS-APIRequest': 'true',
        'Content-Type': 'application/x-www-form-urlencoded'
    }
    auth = HTTPBasicAuth(app_user, app_pw)

    #open excel-file
    wb = openpyxl.load_workbook(file)
    ws = wb['create']

    for row in ws.iter_rows(min_row=2):
        payload = {
            'userid':       row[0].value,
            'password':     row[1].value,
            'email':        row[2].value,
            'displayName':  row[3].value,
            'groups[]':     [x.strip() for x in row[4].value.split(',')],
            'quota':        "0 B",
            'language':     "de"
        }
        r = requests.post(nextcloudURL+'/ocs/v1.php/cloud/users', data=payload, headers=headers, auth=auth)
        print(r)
        print(r.text)
    


def parse_arguments(nc_url):
    """This is where the parsing of commandline arguments is defined.
    The --help helptext is generated implictly by argparse
    """
    parser = argparse.ArgumentParser(description='Create nextcloud-users from excel-file.')
    parser.add_argument('--file', default="users.xlsx", help="name of your excel-file, e.g. 'users.xlsx'")
    parser.add_argument('--nextcloudURL', default=nc_url, help="URL of your Nextcloud, starting with https://")
    parser.add_argument('--overwriteUsers', default=False, type=bool, help="Set to true if you want to overwrite users")
    args = parser.parse_args()
    return args

def main():
    load_dotenv()
    APP_USER = os.getenv("APP_USER")
    APP_PW = os.getenv("APP_PW")
    NC_URL = os.getenv("NC_URL")

    args = parse_arguments(nc_url=NC_URL)

    print('Creating users...')
    create_users_from_file(
        file=args.file,
        nextcloudURL=args.nextcloudURL,
        overwriteUsers=args.overwriteUsers,
        app_user=APP_USER,
        app_pw=APP_PW
    )


if __name__ == "__main__":
    main()
